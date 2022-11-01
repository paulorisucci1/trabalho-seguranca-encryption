import openpyxl


class ExcelGenerator:
    def __init__(self):
        self.__book = openpyxl.Workbook()
        self.__page = self.__book.create_sheet('Métodos de criptografia')
        self.__page.append(['Async encrypt time', 'Async decrypt time', 'Sync encrypt time', 'Sync decrypt time'])
        self.__book.save('Metodos_de_criptografia.xlsx')

    def __create_sheet(self, sheet_name):
        self.__book.create_sheet(sheet_name)

    def recieve_times(self, times):
        self.__page.append(times)
        self.__save_page()

    def __save_page(self):
        self.__book.save('Metodos_de_criptografia.xlsx')

if __name__ == '__main__':
    excel = ExcelGenerator()